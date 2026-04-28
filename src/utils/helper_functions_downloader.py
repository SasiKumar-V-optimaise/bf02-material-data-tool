from config.loader import load_config
from datetime import datetime, timedelta, time
from influxdb_client import InfluxDBClient
from pathlib import Path
import os
import sys
import logging
import time
from datetime import time as dt_time
import tempfile

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import StaleElementReferenceException
import re
import json
import yaml
from ruamel.yaml import YAML
from calendar import month_abbr
import unicodedata
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, date, timedelta



log = logging.getLogger("root")
project_root = Path(__file__).resolve().parents[2]

# Load config
config = load_config()
fixed_order = config.get("FIXED_COLUMN_ORDER", [])





def extract_datetime_from_filename(filename: str) -> datetime:
    """
    Extracts a datetime object from the filename, assuming the format is %Y_%m_%d_%H_%M_%S.csv.
    
    Parameters:
        filename (str): The filename to extract the datetime from.

    Returns:
        datetime: The parsed datetime object from the filename.
    """
    stem = filename.split('.')[0]
    
    # Parse the datetime string assuming format %Y_%m_%d_%H_%M_%S
    try:
        file_datetime = datetime.strptime(stem, "%Y_%m_%d_%H_%M_%S")
        return file_datetime
    except ValueError as e:
        raise ValueError(f"Filename does not match the expected datetime format: {stem}") from e


def setup_browser_driver():
    chrome_options = Options()

    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")

    #  Force isolated environment
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument("--disable-software-rasterizer")

    # TEMP PROFILE (your part, keep it)
    user_data_dir = tempfile.mkdtemp()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")

    service = Service("/usr/bin/chromedriver")

    driver = webdriver.Chrome(service=service, options=chrome_options)

    return driver

def login_eml(driver, wait, LOGIN_URL, USER, PASSWD):
    """
    Log in to the EML web interface using credentials from environment variables.
    Maximizes the browser window, navigates to the login page, and submits the login form.

    Parameters:
        driver: Selenium WebDriver instance.
        wait: WebDriverWait instance for waiting on elements.
    """
    driver.maximize_window()
    print(" Navigating to EML login...")
    driver.get(LOGIN_URL)
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    try:
        username_input = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='text']")))
        password_input = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='password']")))
        username_input.clear()
        username_input.send_keys(USER)
        password_input.clear()
        password_input.send_keys(PASSWD + Keys.ENTER)
        print(" EML login submitted.")
        time.sleep(5)
    except Exception as e:
        print(f" Login skipped or already logged in: {e}")

METADATA_FILE = "downloaded_metadata.json"

def load_metadata():
    if os.path.exists(METADATA_FILE):
        with open(METADATA_FILE, "r") as f:
            return json.load(f)
    return {}

def save_metadata(metadata):
    with open(METADATA_FILE, "w") as f:
        json.dump(metadata, f, indent=2)

def parse_datetime(date_str):
    if not date_str or not isinstance(date_str, str):
        return None
    for fmt in ("%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except Exception:
            continue
    return None

def wait_for_download(download_dir, started_at, timeout=180, stable_checks=2):
    import os, time
    end = time.time() + timeout
    exts = (".crdownload", ".part", ".tmp", ".download")
    last_path, last_size, stable = None, -1, 0

    while time.time() < end:
        # newest non-temp file modified after the click
        files = [os.path.join(download_dir, f) for f in os.listdir(download_dir)]
        files = [p for p in files if os.path.isfile(p) and not p.lower().endswith(exts)]
        cand = max((p for p in files if os.path.getmtime(p) >= started_at), default=None, key=os.path.getmtime)

        if cand:
            size = os.path.getsize(cand)
            if cand == last_path and size == last_size:
                stable += 1
                if stable >= stable_checks:
                    return cand
            else:
                last_path, last_size, stable = cand, size, 0
        time.sleep(1)
    return None

def go_to_file_station_and_download(driver, wait, target_files, ROOT_URL, HOURLY_URL, selected_modes, run_date, target_filename=None):
    norm = lambda s: re.sub(r"\s+", " ", (s or "")).strip().lower()
    is_today_mode = "--today" in sys.argv
    meta = load_metadata()
    skipped, download_dir = set(), os.path.expanduser("~/Downloads")

    def js_rows():
        return driver.execute_script("""
            return Array.from(document.querySelectorAll('.x-grid3-body .x-grid3-row')).map(row=>{
              const t=[...row.querySelectorAll('.x-grid3-cell-inner')].map(c=>c.innerText.trim());
              return {el:row,name:t[0]||"",size:t[1]||"",type:t[2]||"",modified:t[3]||""};
            });
        """)

    def cleanup_prefix(prefix, keep_name):
        for k in [k for k in list(meta) if k.startswith(prefix) and k != keep_name]:
            del meta[k]

    def should_download(name, modified_str):
        cur_dt = parse_datetime(modified_str); prev_dt = parse_datetime(meta.get(name))
        return (cur_dt, prev_dt, (not prev_dt or cur_dt != prev_dt))

    # ---------- ROOT (RM / DPR / HOT_METAL / RM & HM) ----------
    mode_map = {
        "rm": "11A BF-02 BUNKER",
        "dpr": "BF-02 DPR",
        "hot_metal": "06 BF-02- HOT METAL, SLAG & GAS",
        "rm_hm": "RM & HM"   # new file type
    }

    driver.get(ROOT_URL)
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(5)

    try:
        wait.until(lambda d: d.find_elements(By.CLASS_NAME, "x-grid3-row"))
    except:
        logger.info(" File list did not appear.")
        return skipped

    rows = js_rows()

    for mode in [m for m in ["rm", "dpr", "hot_metal", "rm_hm"] if m in selected_modes]:
        want = mode_map[mode]
        row = next((r for r in rows if norm(want) in norm(r["name"])), None)
        if not row:
            logger.info(f" '{want}' not found.")
            skipped.add(want)
            continue

        cur_dt, prev_dt, need = should_download(row["name"], row["modified"])
        if not need:
            logger.info(f" No update for '{row['name']}'")
            skipped.add(want)
            continue

        logger.info(f" Downloading {row['name']}…")
        start_ts = time.time()
        ActionChains(driver).move_to_element(row["el"]).double_click(row["el"]).perform()
        downloaded = wait_for_download(download_dir, started_at=start_ts, timeout=240, stable_checks=2)
        logger.info(f"{'Done' if downloaded else 'Download failed'}: {row['name']}")

        # cleanup previous entries of same type
        if mode == "dpr":
            cleanup_prefix("BF-02 DPR", row["name"])
        elif mode == "hot_metal":
            cleanup_prefix("06 BF-02- HOT METAL", row["name"])
        elif mode == "rm":
            cleanup_prefix("11A BF-02 BUNKER", row["name"])
        elif mode == "rm_hm":
            cleanup_prefix("RM & HM", row["name"])  #  added cleanup for new type

        meta[row["name"]] = (cur_dt or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
        save_metadata(meta)

    # ---------- HOURLY (Charge & Dump) ----------
    if "charge" in selected_modes:
        driver.get(HOURLY_URL); wait.until(lambda d: d.execute_script("return document.readyState")=="complete"); time.sleep(1)
        try: panel = wait.until(EC.presence_of_element_located((By.CLASS_NAME,"x-grid3-scroller")))
        except: print(" Could not locate scroll panel."); skipped.add("charge_and_dump"); return skipped

        dt = datetime.today() if is_today_mode else datetime.strptime(run_date, "%d-%b-%Y")
        stems = [f"CHARGE_AND_DUMP_REPORT_{dt.day}_{dt.month}_{dt.year}", f"CHARGE_AND_DUMP_REPORT_{dt.day:02d}_{dt.month:02d}_{dt.year}"]
        candidates = {*(s+".xls" for s in stems), *(s+".xlsx" for s in stems)}
        logger.info(f" Looking for: {', '.join(sorted(candidates))}")

        prev = meta.get("HOURLY_REPORT", {}) if is_today_mode else {}
        seen, found = set(), False

        for _ in range(60):
            rows = driver.execute_script("""
            return [...document.querySelectorAll('.x-grid3-body .x-grid3-row')].map(r=>{
                const t=[...r.querySelectorAll('.x-grid3-cell-inner')].map(c=>c.innerText.trim());
                return {el:r,n:t[0]||'',m:t[3]||''};
            });
            """)
            for r in rows:
                n = r["n"].strip()
                if not n or n in seen: continue
                seen.add(n)
                if not (n in candidates or any(n.startswith(s) and n.lower().endswith(('.xls','.xlsx')) for s in stems)): continue

                mod = parse_datetime(r["m"]) or datetime.now()
                if is_today_mode and isinstance(prev, dict) and prev.get("name")==n and parse_datetime(prev.get("modified") or "")>=mod:
                    logger.info(f" No update in HOURLY file since {prev.get('modified')}"); skipped.add("charge_and_dump"); return skipped

                driver.execute_script("arguments[0].scrollIntoView({block:'center'})", r["el"])
                ActionChains(driver).move_to_element(r["el"]).double_click(r["el"]).perform()
                _ = wait_for_download(os.path.expanduser('~/Downloads'), time.time(), 240, 2)
                if is_today_mode:
                    meta["HOURLY_REPORT"]={"name":n,"modified":mod.strftime("%Y-%m-%d %H:%M:%S")}
                    save_metadata(meta)
                logger.info(f" Downloaded hourly: {n}"); found=True; break
            if found: break
            ActionChains(driver).move_to_element(panel).click().send_keys(Keys.PAGE_DOWN).perform(); time.sleep(0.6)

        if not found: logger.warning(" Hourly file not found."); skipped.add("charge_and_dump")

    return skipped


def parse_date_input(start_date):
    if isinstance(start_date, list):
        return [datetime.strptime(d, "%d-%b-%Y").date() for d in start_date]
    return [datetime.strptime(start_date, "%d-%b-%Y").date()]

def normalize_columns(df):
    df.columns = df.columns.str.strip().str.replace(" ", "_").str.upper()
    return df



def average_numeric_group(df, group_cols, skip_cols=[], preserve_order_col="MERGE_KEY"):
    """
    Groups by group_cols, averages numeric columns (ignoring 0s and NaNs),
    discards non-numeric values, and preserves row order.
    """
    results = []

    # Preserve original order
    order_map = (
        df.drop_duplicates(preserve_order_col)
        .reset_index()
        .set_index(preserve_order_col)["index"]
        .to_dict()
    )

    for keys, group in df.groupby(group_cols):
        group = group.drop(columns=skip_cols, errors="ignore")

        avg_row = {}

        # Identify numeric columns reliably
        numeric_cols = []
        for col in group.columns:
            if col in group_cols:
                continue

            # Attempt to convert entire column to numeric (coerce errors to NaN)
            converted = pd.to_numeric(group[col], errors="coerce")
            valid_ratio = converted.notna().sum() / len(group)

            if valid_ratio >= 0.5:
                # Keep column as numeric
                numeric_cols.append(col)

                # Treat 0s as missing
                cleaned = converted.mask(converted == 0, pd.NA)

                # Average ignoring 0 and NaNs
                avg_val = cleaned.mean(skipna=True)
                if pd.notna(avg_val):
                    avg_row[col] = avg_val

        # Add back group keys
        if isinstance(keys, tuple):
            for i, col in enumerate(group_cols):
                avg_row[col] = keys[i]
            merge_key = f"{keys[0]}_{keys[1]}"
        else:
            avg_row[group_cols[0]] = keys
            merge_key = str(keys)

        avg_row["_ORDER"] = order_map.get(merge_key, -1)
        avg_row[preserve_order_col] = merge_key
        results.append(avg_row)

    if not results:
        return pd.DataFrame()

    df_out = pd.DataFrame(results)
    df_out.sort_values("_ORDER", inplace=True)
    df_out.drop(columns=["_ORDER"], inplace=True)
    return df_out

def average_shift_blocks(df):
    """
    Averages rows per (DATE, SHIFT_GROUP), skipping zero/NaNs for numeric cols,
    merges strings, and preserves original SHIFT order.
    """
    df["SHIFT_GROUP"] = df["SHIFT"].str[0].str.upper()
    df["MERGE_KEY"] = df["DATE"].astype(str) + "_" + df["SHIFT_GROUP"]

    # Preserve first-seen MERGE_KEY order
    shift_order = df.drop_duplicates("MERGE_KEY")["MERGE_KEY"].tolist()

    skip_cols = [
        col for col in df.columns if col.upper() in {"SHIFT", "SHIFT_GROUP"}
    ]

    avg_df = average_numeric_group(
        df,
        group_cols=["DATE", "SHIFT_GROUP"],
        skip_cols=skip_cols,
        preserve_order_col="MERGE_KEY"
    )

    if avg_df.empty:
        return pd.DataFrame(columns=["DATE", "SHIFT"])

    # Rename SHIFT_GROUP to SHIFT
    avg_df.rename(columns={"SHIFT_GROUP": "SHIFT"}, inplace=True)

    # Restore original shift order using preserved list
    avg_df["MERGE_KEY"] = avg_df["DATE"].astype(str) + "_" + avg_df["SHIFT"]
    avg_df["_ORDER"] = avg_df["MERGE_KEY"].apply(lambda x: shift_order.index(x) if x in shift_order else -1)
    avg_df = avg_df.sort_values("_ORDER").drop(columns=["MERGE_KEY", "_ORDER"]).reset_index(drop=True)

    # Reorder columns to put DATE and SHIFT first
    cols = ["DATE", "SHIFT"] + [col for col in avg_df.columns if col not in {"DATE", "SHIFT"}]
    avg_df = avg_df[cols]

    logger.info(" AVG is Done")
    return avg_df


def read_excel_sheet(xls, sheet, cols, hdr, start_date=None, output_dir="outputs"):
    if sheet not in xls.sheet_names:
        logger.info(f" Sheet '{sheet}' missing, skipping.")
        return None

    df = pd.read_excel(xls, sheet_name=sheet, usecols=cols, header=hdr).dropna(how="all").reset_index(drop=True)

    # Normalize column names (optional)
    df.columns = [str(col).strip().upper() for col in df.columns]

    # Define always-keep-as-text columns
    always_text_cols = ['DATE', 'SHIFT', 'SOURCE', 'BUNKER','BNK NO', 'ONLINE/OFFLINE','SORUCE']

    for col in df.columns:
        if any(key in col for key in always_text_cols):
            continue
        # Try to coerce to float; replace non-numeric with NaN
        df[col] = pd.to_numeric(df[col], errors='coerce')

    return df.drop(columns=["TIME"], errors="ignore")

def filter_by_date_and_shift(df, date_list, sheet_name=None, logger=None):
    col_map = {col.strip().upper(): col for col in df.columns}
    date_col = col_map.get("DATE")
    shift_col = next((col_map[k] for k in col_map if "SHIFT" in k), None)

    #  Exception for SP-02 (RI-RDI) sheet (ignore shift validation)
    if sheet_name == "SP-02 (RI-RDI)":
        if not date_col:
            msg = "  Missing DATE column — skipping this sheet"
            print(msg, df) if logger is None else logger.warning(msg)
            return None
        parsed = pd.to_datetime(df[date_col], format="%d-%m-%Y", errors="coerce")
        if parsed.isna().all():
            parsed = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
        df[date_col] = parsed.dt.date
        df = df[df[date_col].isin(date_list)].reset_index(drop=True)
        return df

    # 🔹 Normal case (DATE + SHIFT required)
    if not date_col or not shift_col:
        msg = "  Missing DATE or SHIFT column — skipping this sheet"
        print(msg, df) if logger is None else logger.warning(msg)
        return None

    parsed = pd.to_datetime(df[date_col], format="%d-%m-%Y", errors="coerce")
    if parsed.isna().all():
        parsed = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    df[date_col] = parsed.dt.date
    df[shift_col] = df[shift_col].astype(str).str.strip().str.upper()
    VALID_SHIFTS = {"A", "B", "C"}
    df[shift_col] = df[shift_col].apply(
        lambda x: x[0] if x and isinstance(x, str) and x[0] in VALID_SHIFTS else pd.NA
    )
    before = len(df)
    df = df[df[date_col].isin(date_list) & df[shift_col].notna()].reset_index(drop=True)
    after = len(df)
    # msg = f"   ↪  Kept {after}/{before} rows with valid DATE and SHIFT in {VALID_SHIFTS}"
    # print(msg) if logger is None else logger.info(msg)
    return df


def split_online_offline_and_merge(df):
    df = normalize_columns(df)
    df["ONLINE/OFFLINE"] = df["ONLINE/OFFLINE"].astype(str).str.strip()
    df["SHIFT"] = df["SHIFT"].astype(str).str.strip()
    df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce", dayfirst=True).dt.date
    if "BNK_NO" in df.columns:
        df["BNK_NO"] = df["BNK_NO"].astype(str).str.strip()
    df["MERGE_KEY"] = df["DATE"].astype(str) + "_" + df["SHIFT"]
    df["IS_ONLINE"] = df["ONLINE/OFFLINE"].str.contains(r"(?:ONLINE|EML)", case=False, na=False)

    def reduce(g, suffix):
        base = {col: g[col].iloc[0] for col in ["MERGE_KEY", "DATE", "SHIFT"] if col in g.columns}
        numeric_cols = []
        num_data = {}
        for col in g.columns:
            if col in base or col in ["IS_ONLINE", "MERGE_KEY", "DATE", "SHIFT"]:
                continue
            try:
                converted = pd.to_numeric(g[col], errors="coerce")
                if converted.notna().sum() >= 0.8 * len(g):
                    vals = converted.replace(0, pd.NA).dropna()
                    num_data[col + suffix] = round(vals.mean(skipna=True), 3) if not vals.empty else pd.NA
                    numeric_cols.append(col)
            except:
                continue
        for col in g.columns:
            if col in base or col in numeric_cols or col in ["IS_ONLINE"]:
                continue
            unique_vals = g[col].dropna().astype(str).str.strip().unique()
            base[col + suffix] = " | ".join(sorted(set(unique_vals)))
        base.update(num_data)
        return base

    records = []
    for (key, shift), group in df.groupby(["MERGE_KEY", "SHIFT"]):
        online, offline = group[group["IS_ONLINE"]], group[~group["IS_ONLINE"]]
        if len(online) > 1:
            logger.info(f"\n Multiple ONLINE rows for DATE: {group['DATE'].iloc[0]}, SHIFT: {shift}")
            logger.info(online.to_string(index=False))
        if len(offline) > 1:
            logger.info(f"\n  Multiple OFFLINE rows for DATE: {group['DATE'].iloc[0]}, SHIFT: {shift}")
            logger.info(offline.to_string(index=False))
        combined = {}
        if not online.empty: combined.update(reduce(online, "_ON"))
        if not offline.empty: combined.update(reduce(offline, "_OFF"))
        records.append(combined)

    df_out = pd.DataFrame(records)
    final_df = pd.merge(df.drop_duplicates("MERGE_KEY")[["MERGE_KEY"]].reset_index(), df_out, on="MERGE_KEY")
    return final_df.drop(columns=["MERGE_KEY", "index"])

def prefix_columns(df, prefix):
    df.columns = [prefix + str(c) for c in df.columns]
    return df

def write_combined_file(parts, combined_path, date_list):

    SHIFT_ORDER = ["C", "A", "B"]
    SHIFT_TIME  = {"A": "07:00", "B": "15:00", "C": "23:00"}

    combined = pd.concat(parts, axis=1)

    # pick first *_DATE col
    date_col = next((c for c in combined.columns if c.upper().endswith("_DATE")), None)
    if not date_col:
        raise ValueError("No _DATE columns found")
    combined["Date"] = pd.to_datetime(combined[date_col], errors="coerce")
    combined.drop(columns=[c for c in combined.columns if c.upper().endswith("_DATE")],
                  inplace=True, errors="ignore")

    # assign shifts cyclically
    combined["SHIFT"] = [SHIFT_ORDER[i % 3] for i in range(len(combined))]

    # build datetime string for all rows
    shift_times = combined["SHIFT"].map(SHIFT_TIME)
    combined["Date"] = pd.to_datetime(combined["Date"].dt.strftime("%Y-%m-%d") + " " + shift_times)

    # subtract 1 day for C-shift
    combined.loc[combined["SHIFT"] == "C", "Date"] -= pd.Timedelta(days=1)

    combined.drop(columns="SHIFT", inplace=True)
    combined = combined[["Date"] + [c for c in combined.columns if c != "Date"]]

    # avoid duplicates if appending multiple days
    if len(date_list) > 1 and os.path.exists(combined_path):
        existing = pd.read_excel(combined_path)
        existing["Date"] = pd.to_datetime(existing["Date"], errors="coerce")
        combined = pd.concat(
            [existing[~existing["Date"].dt.date.isin(pd.to_datetime(date_list).date)], combined],
            ignore_index=True
        )

    combined.to_excel(combined_path, index=False)

def read_rm_sheet(file_path, RM_SHEET_CONFIG, start_date="11-Jul-2025", output_dir="outputs"):
    logger.info(f"\n  Reading Excel file: {file_path}")
    date_list = parse_date_input(start_date)
    logger.info(f" Including rows from {date_list[0]} to {date_list[-1]}" if len(date_list) > 1 else f"   Including rows with DATE == {date_list[0]}")
    os.makedirs(output_dir, exist_ok=True)
    combined_path = os.path.join(output_dir, "combined_bunker_data.xlsx")
    xls = pd.ExcelFile(file_path)
    parts = []

    for key, cfg in RM_SHEET_CONFIG.items():
        sheet, cols, hdr = cfg["sheet_name"], cfg["columns"], cfg["header_row"] - 1
        prefix = cfg.get("col_prefix", "")
        df = read_excel_sheet(xls, sheet, cols, hdr)
        if df is None:
            continue

        df = normalize_columns(df)
        logger.info(sheet)
        df = filter_by_date_and_shift(df, date_list,sheet_name=sheet )
        if df is None or df.empty:
            logger.warning(f"   {key}: no valid data")
            
            continue

        if "ONLINE/OFFLINE" in df.columns:
            df = split_online_offline_and_merge(df)
            # logger.info(f"     {key}: merged ONLINE + OFFLINE per SHIFT+DATE")

        # Apply averaging if multiple rows for a normalized shift group
        if "SHIFT" in df.columns and "DATE" in df.columns:
            shift_counts = df.groupby(["DATE", "SHIFT"]).size().reset_index(name="count")
            if any(shift_counts["count"] > 1):
                df = average_shift_blocks(df)
                
                logger.info(f"  {key}: averaged multiple rows per SHIFT block")

        df = prefix_columns(df, prefix)
        parts.append(df)


    if parts:
        write_combined_file(parts, combined_path, date_list)
    else:
        logger.warning("No valid data combined — exiting.")




# Function to parse sheet date from name like "Jun'25"
def parse_sheet_date(name):
    if not name:
        return None

    # normalize: trim spaces
    name = name.strip()

    # allow any case, allow trailing spaces
    m = re.match(r"^([A-Za-z]+)\s*'\s*(\d{2})$", name, re.IGNORECASE)
    if not m:
        return None

    try:
        month_str = m.group(1)[:3].title()   # JAN, jan → Jan
        year = 2000 + int(m.group(2))
        month = list(month_abbr).index(month_str)
        return (year, month)
    except Exception:
        return None


def update_dpr_config_from_excel(excel_path, yaml_path, run_date):


    config = load_config(yaml_path)
    yaml = YAML()
    yaml.preserve_quotes = True

    sheets = config.get("DPR_CONFIG", {}).get("sheets", {})
    if not sheets:
        raise RuntimeError("Missing DPR_CONFIG.sheets in YAML")

    # convert run_date to datetime.date
    run_dt = datetime.strptime(run_date, "%d-%b-%Y").date()
    target_sheet = None
    target_dt = (run_dt.year, run_dt.month)

    # match sheet by run date month and year
    wb = load_workbook(excel_path, data_only=True)
    dated = [(s, parse_sheet_date(s)) for s in wb.sheetnames if parse_sheet_date(s)]

    for sname, parsed in dated:
        if parsed and (parsed[0], parsed[1]) == target_dt:
            target_sheet = sname
            break

    if not target_sheet:
        print(f" No matching sheet found for run date {run_date}. Using latest available sheet.")
        target_sheet, _ = max(dated, key=lambda x: x[1])

    sheet_key = target_sheet.replace("'", "")
    ws = wb[target_sheet]

    if sheet_key not in sheets:
        old_key = next(iter(sheets))
        sheets[sheet_key] = sheets.pop(old_key)
        print(f" Renamed YAML month block '{old_key}' ➔ '{sheet_key}'")

    block = sheets[sheet_key]
    block["sheet_name"] = target_sheet

    old_rows = block.get("rows", {})
    found_rows = {}

    for r in range(1, ws.max_row + 1):
        texts = [str(ws.cell(r, c).value).strip() for c in range(1, 8) if ws.cell(r, c).value]
        for label in old_rows:
            if label in texts and label not in found_rows:
                found_rows[label] = r

    block["rows"] = {}

    for k in old_rows:
        if k in found_rows:
            block["rows"][k] = found_rows[k]
        else:
            block["rows"][k] = 0  # ← Retain but empty
            print(f" '{k}' not found in sheet '{target_sheet}' → keeping key with None value")


    with open(yaml_path, "w") as f:
        yaml.dump(config, f)

    print(f" Updated '{yaml_path}' with sheet '{target_sheet}' → rows: {block['rows']}")


def read_dpr_sheet(
    file_path: str,
    config: dict,
    start_date: str = "11-Jun-2025",
    output_dir: str = "outputs"
):
    """
    Reads DPR sheets based on YAML config, renames fields, filters by start_date,
    and saves combined output to a single fixed file: combined_dpr_data.xlsx
    """
 

    dpr_sheets = config["DPR_CONFIG"]["sheets"]
    os.makedirs(output_dir, exist_ok=True)
    start_dt = datetime.strptime(start_date, "%d-%b-%Y").date()
    wb = load_workbook(file_path, data_only=True)

    all_parts = []

    logger.info(f"\n Reading DPR Excel file: {file_path}")
    logger.info(f" Filtering for date: {start_dt}")

    for sheet_key, cfg in dpr_sheets.items():
        sheet_name = cfg["sheet_name"]
        date_row = cfg["date_row"] - 1
        col_start, col_end = cfg["date_cols"]
        col_range = range(
            column_index_from_string(col_start) - 1,
            column_index_from_string(col_end)
        )

        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found — skipping.")
            continue

        ws = wb[sheet_name]

        # Step 1: Read date headers
        raw_dates = [ws.cell(row=date_row + 1, column=col + 1).value for col in col_range]
        parsed_dates = []
        valid_indices = []

        for i, val in enumerate(raw_dates):
            parsed = pd.to_datetime(val, errors="coerce")
            if pd.notna(parsed):
                parsed_dates.append(parsed.date())
                valid_indices.append(i)

        if not parsed_dates:
            print(f" No valid dates found in sheet '{sheet_name}' — skipping.")
            continue

        # Step 2: Read rows for this sheet
        raw_data = {}
        for label, row in cfg["rows"].items():
            if not row or row < 1:
                print(f" Skipping '{label}' — invalid row index: {row}")
                continue
            all_values = [ws.cell(row=row, column=col + 1).value for col in col_range]
            filtered_values = [all_values[i] for i in valid_indices]
            raw_data[label] = filtered_values
        # Step 3: Drop columns that are empty across all rows
        non_empty_mask = [
            any(raw_data[label][i] is not None for label in raw_data)
            for i in range(len(valid_indices))
        ]
        filtered_dates = [parsed_dates[i] for i, keep in enumerate(non_empty_mask) if keep]

        if not filtered_dates:
            print(f" No data found for sheet '{sheet_name}' on any date — skipping.")
            continue

        # Step 4: Build DataFrame for this sheet
        df = pd.DataFrame({"Date": filtered_dates})
        rename_map = cfg.get("rename_map", {})
        reverse_map = {v[0]: k for k, v in rename_map.items() if isinstance(v, list) and len(v) == 1}

        for original, all_values in raw_data.items():
            values = [all_values[i] for i, keep in enumerate(non_empty_mask) if keep]
            colname = reverse_map.get(original, original)
            df[colname] = values

        def compute_total_coke(row):
            val1 = row.get("COKE_ONLINE_MT", 0)
            val2 = row.get("COKE_OFFLINE_MT", 0)
            val3 = row.get("TOTAL_PURCHASE_COKE_MT", 0)

            val1 = 0 if pd.isna(val1) or val1 == 0 else val1
            val2 = 0 if pd.isna(val2) or val2 == 0 else val2
            val3 = 0 if pd.isna(val3) or val3 == 0 else val3

            return val1 + val2 + val3

        # Derive TOTAL_COKE_MT as sum of COKE_ONLINE_MT and COKE_OFFLINE_MT
        if "COKE_ONLINE_MT" in df.columns or "COKE_OFFLINE_MT" in df.columns or "TOTAL_PURCHASE_COKE_MT" in df.columns:
            df["TOTAL_COKE_MT"] = df.apply(compute_total_coke, axis=1)


        # Step 5: Filter for only start_date
        before = len(df)
        df = df[df["Date"] == start_dt].reset_index(drop=True)
        after = len(df)

        print(f" Sheet '{sheet_name}': kept {after}/{before} rows for {start_dt}")

        if not df.empty:
            # df.insert(0, "Sheet", sheet_key)
            all_parts.append(df)

    # Final save
    if not all_parts:
        print(" No DPR data found — nothing to write.")
        return

    final_df = pd.concat(all_parts, ignore_index=True)
    os.path.join(output_dir, "combined_dpr_data.xlsx")
    print(f"\n DPR data extracted for {start_date}")
    return final_df




def update_hot_metal_config_from_excel(file_path: str, hot_metal_path: str, run_date: str):
    mon = datetime.strptime(run_date.strip(), "%d-%b-%Y").strftime("%b").upper()
    yy  = datetime.strptime(run_date.strip(), "%d-%b-%Y").strftime("%y")
    wb  = load_workbook(file_path, read_only=True)
    try:
        target = next(s for s in wb.sheetnames if mon in s.upper() and yy in s)
    except StopIteration:
        raise ValueError(f"No matching sheet for {mon}-{yy}. Found: {wb.sheetnames}")

    try:
        with open(hot_metal_path, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}
    except FileNotFoundError:
        cfg = {}

    hm = cfg.setdefault("HOT_METAL_CONFIG", {})
    sheets = hm.setdefault("sheets", {})
    if target not in sheets and sheets:
        sheets[target] = sheets.pop(next(iter(sheets)))  # reuse structure
    hm["sheet_name"] = target

    with open(hot_metal_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(cfg, f, sort_keys=False)
    print(f" HOT_METAL_CONFIG updated in '{hot_metal_path}' → sheet_name: {target}")


def read_hot_metal_sheet(file_path, start_date, config, output_dir=None):

    # normalize input dates → list[date]
    def _to_date(d):
        if isinstance(d, date): return d
        if isinstance(d, datetime): return d.date()
        return datetime.strptime(str(d).strip(), "%d-%b-%Y").date()
    fdates = [_to_date(d) for d in (start_date if isinstance(start_date, list) else [start_date])]

    # pull HM config (from hot_metal.yaml)
    hm = (config or {}).get("HOT_METAL_CONFIG", {})
    mon_key = (hm.get("sheet_name") or "").strip()          # e.g. "AUG-25"
    block   = (hm.get("sheets")   or {}).get(mon_key, {})   # block for that month
    usecols = block.get("columns", "A:Z")
    hdr     = sorted(block.get("header_row", [3, 4]))
    top     = max(hdr)

    # resolve actual Excel sheet: exact match or by month token + yy
    xls = pd.ExcelFile(file_path)
    sheet = mon_key if mon_key in xls.sheet_names else None
    if not sheet and "-" in mon_key:
        mon, yy = mon_key.split("-", 1)
        mon, yy = mon.strip().upper(), yy.strip()
        sheet = next((s for s in xls.sheet_names if mon in s.upper() and yy in s), None)
    if not sheet:
        raise ValueError(f"Sheet for '{mon_key}' not found. Available: {xls.sheet_names}")

    # build 2-line header

    H   = xls.parse(sheet, header=None, usecols=usecols, nrows=top + 1).fillna("")
    h1  = H.iloc[hdr[0]].astype(str).str.strip()
    h2  = H.iloc[hdr[1]].astype(str).str.strip()
    cols = [(a if a and not b else b if b and not a else f"{a} | {b}" if (a or b) else "")
            for a, b in zip(h1, h2)]
    cols = [c or f"COL_{i+1}" for i, c in enumerate(cols)]
    seen = {}
    cols = [c if not seen.setdefault(c, 0) else f"{c}_{seen.update({c: seen[c]+1}) or seen[c]}" for c in cols]

    # read data rows
    df = xls.parse(sheet, header=None, usecols=usecols, skiprows=top + 1).dropna(how="all")
    df.columns = cols


    # locate date/time columns
    date_col = next((c for c in df.columns if "DATE" in str(c).upper()), None)
    time_col = next((c for c in df.columns if "RECD TIME" in str(c).upper()), None)

    if date_col is None:
        raise ValueError(f" 'DATE' column not found. First columns: {df.columns[:10].tolist()}")

    # convert to datetime
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    df = df[df[date_col].notna()].copy()



    # # merge DATE + RECD TIME → 'date'
    # if time_col in df.columns:
    #     # Hot-metal tends to use HH.MM format; keep tolerant parsing
    #     t = pd.to_datetime(df[time_col].astype(str), format="%H.%M", errors="coerce").dt.time
    #     df[time_col] = t
    #     df["date"] = df.apply(
    #         lambda r: datetime.combine(r[date_col].date(), r[time_col]) - timedelta(minutes=16)
    #         if pd.notna(r[time_col]) else r[date_col],
    #         axis=1
    #     )
    # else:
    #     df["date"] = df[date_col]

    if time_col in df.columns:

        def parse_hm_time(value):
            if pd.isna(value):
                return None

            # Normalize numeric values like 0.4 → "0.40"
            value = f"{float(value):.2f}"

            # Split hour + minutes
            h, m = value.split(".")
            h, m = int(h), int(m)

            # Fix invalid minutes like 60, 75, etc.
            if m >= 60:
                extra_hour = m // 60
                m = m % 60
                h += extra_hour

            # Prevent hour overflow (wrap to next day)
            h = h % 24

            return dt_time(h, m)
        # Parse time correctly
        df[time_col] = df[time_col].apply(parse_hm_time)

        # Build final datetime and subtract 16 minutes
        df["date"] = df.apply(
            lambda row: datetime.combine(row[date_col].date(), row[time_col]) - timedelta(minutes=16)
            if row[time_col] else row[date_col],
            axis=1
        )

    else:
        df["date"] = df[date_col]

    # filter requested dates and format
    df = df[df["date"].dt.date.isin(fdates)].copy()
    df["date"] = df["date"].dt.strftime("%Y-%m-%d %H:%M")

    if output_dir:

        os.makedirs(output_dir, exist_ok=True)
        out_path = os.path.join(output_dir, "filtered_hotmetal_data.xlsx")
        df.to_excel(out_path, index=False)
        logger.info(f"Filtered Hot Metal data written to {out_path}")

    return df





import psycopg2
logger = logging.getLogger(__name__)


class ChargeDataProcessor:
    """
    Process hopper-level charge data into hourly material-level data.
    Material mapping comes dynamically from Neon DB for each hour.
    PCI_MT is fetched from InfluxDB and merged into the final report.
    """
    config = load_config(os.path.join("src", "config", "setting.yaml"))
    def __init__(self, file_today, file_yesterday, output_dir, run_date_str,
                 neon_cfg, influx_cfg=None, material_groups=None):

        self.files = [file_yesterday, file_today]
        self.output_dir = output_dir

        # Parse date like "02-Nov-2025"
        self.target_date = datetime.strptime(
            re.search(r"(\d{1,2})-(\w+)-(\d{4})", run_date_str).group(0),
            "%d-%b-%Y"
        )

        self.neon_cfg = neon_cfg
        self.influx_cfg = influx_cfg
        self.material_groups = material_groups or {}

    # ---------------------------
    #  Neon DB Mapping Fetcher
    # ---------------------------
    def _fetch_material_mapping_for_hour(self, dt):
        try:
            conn = psycopg2.connect(**self.neon_cfg)
            cur = conn.cursor()

            query = """
                SELECT DISTINCT ON (hopper)
                    hopper, material
                FROM hopper_material_history
                WHERE valid_from <= %s
                AND (valid_upto IS NULL OR valid_upto > %s)
                ORDER BY hopper, valid_from DESC;
            """

            cur.execute(query, (dt, dt))
            rows = cur.fetchall()

            cur.close()
            conn.close()

            mapping = {}

            for hopper_name, material_name in rows:

                # ---------------------------
                # FIX: Remove leading zeros
                # ---------------------------
                clean_name = hopper_name.upper().replace(" ", "_")

                match = re.search(r'HOPPER_0*(\d+)$', clean_name)
                if match:
                    hopper_num = match.group(1)  # e.g. 01 → 1
                    hopper_col = f"HOPPER_{hopper_num}_ACT"
                else:
                    hopper_col = clean_name + "_ACT"

                # Material normalization
                material_key = re.sub(r'[^a-z0-9_]', '', 
                    re.sub(r"\s+", "_", material_name.lower())
                    .replace("-", "_")
                    .replace("/", "_")
                ) + "_mt"

                mapping.setdefault(material_key, []).append(hopper_col)
            # print(mapping)
            return mapping

        except Exception as e:
            logger.error("Error fetching mappings from Neon DB: %s", e)
            return {}

    # ---------------------------
    #  Excel Loader
    # ---------------------------
    def _load_excel(self, path):
        if not (path and os.path.exists(path)):
            return pd.DataFrame()

        xl = pd.ExcelFile(path)
        sheet = next((s for s in xl.sheet_names if "SH" in s.upper()), None)
        if not sheet:
            return pd.DataFrame()

        df = xl.parse(sheet, skiprows=6)
        df = df.loc[:, ~df.columns.str.contains("Unnamed", case=False)]
        df.drop(columns=[c for c in df.columns if "CHARGE_NO" in c.upper()],
                inplace=True, errors="ignore")
        df["DATETIME"] = pd.to_datetime(df["DATETIME"], errors="coerce")
        return df


    # ---------------------------
    #  PCI MT Fetcher from Influx
    # ---------------------------
    def _get_pci_mt_from_influx(self):
        if not self.influx_cfg:
            logger.warning("InfluxDB config missing. Skipping pci_mt enrichment.")
            return pd.DataFrame()

        start_utc_dt = (self.target_date - timedelta(hours=6, minutes=30))
        stop_utc_dt  = (self.target_date + timedelta(days=1)
                        - timedelta(hours=5, minutes=30))
        print(start_utc_dt, stop_utc_dt)

        start_utc = start_utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        stop_utc  = stop_utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ")

        # q = f"""
        # from(bucket: "bf2_evonith_raw")
        # |> range(start: {start_utc}, stop: {stop_utc})
        # |> filter(fn: (r) => r._measurement == "process_params" and
        #                     (r._field == "coal_rate_actual_value" or r._field == "production_per_hour"))
        # |> aggregateWindow(every: 1h, fn: mean, createEmpty: true, offset: 30m)
        # |> pivot(rowKey:["_time"], columnKey:["_field"], valueColumn:"_value")
        # |> yield(name: "hourly_avg")
        # """
        q = f"""
import "math"

from(bucket: "bf2_evonith_raw")
|> range(start: {start_utc}, stop: {stop_utc})
|> filter(fn: (r) => r._measurement == "process_params" and
    (r._field == "coal_rate_actual_value" or r._field == "production_per_hour"))
|> map(fn: (r) => ({{
    r with _value: math.abs(x: r._value)
}}))
|> aggregateWindow(every: 1h, fn: mean, createEmpty: true, offset: 30m)
|> pivot(rowKey:["_time"], columnKey:["_field"], valueColumn:"_value")
|> yield(name: "hourly_avg")
"""



        try:
            with InfluxDBClient(**self.influx_cfg) as client:
                df = client.query_api().query_data_frame(q)
                if df.empty:
                    return pd.DataFrame()

                df = df.rename(columns={"_time": "DATETIME"})
                df["DATETIME"] = (
                    pd.to_datetime(df["DATETIME"], utc=True)
                    .dt.tz_convert("Asia/Kolkata")
                    .dt.floor("h")
                    .dt.tz_localize(None)
                )

                df["pci_mt"] = (
                    df["coal_rate_actual_value"] *
                    df["production_per_hour"]
                ) / 1000

                mask = (df["DATETIME"] >= self.target_date) & (
                        df["DATETIME"] < self.target_date + timedelta(days=1)
                )
                df = df.loc[mask, ["DATETIME", "pci_mt"]]

                idx = pd.date_range(self.target_date,
                                    self.target_date + timedelta(days=1),
                                    freq="1h",
                                    inclusive="left")

                df = df.set_index("DATETIME").reindex(idx)
                df.index.name = "DATETIME"
                df = df.reset_index()


                return df

        except Exception as e:
            logger.error(f"Failed to fetch pci_mt from InfluxDB: {e}")
            return pd.DataFrame()

    # ---------------------------
    #  MAIN PROCESS METHOD
    # ---------------------------
    def process(self):
        # Load data (yesterday + today)
        data = pd.concat(
            [self._load_excel(p) for p in self.files if p],
            ignore_index=True
        )

        if data.empty or "DATETIME" not in data.columns:
            logger.warning("No charge data available.")
            return None

        # ------------------------------------------------
        # STEP 1: Shift timestamps to represent NEXT hour
        # ------------------------------------------------
        data["DATETIME"] = (
            data["DATETIME"]
            .dt.floor("h") +
            pd.Timedelta(hours=1)
        )

        # ------------------------------------------------
        # STEP 2: Hourly hopper aggregation
        # ------------------------------------------------
        hourly = (
            data
            .groupby("DATETIME")[[c for c in data.columns if "ACT" in c]]
            .sum()
            .reset_index()
        )

        # ------------------------------------------------
        # STEP 3: Keep only target date hours
        # ------------------------------------------------
        start = self.target_date
        end   = self.target_date + timedelta(days=1)

        hourly = hourly[
            (hourly["DATETIME"] >= start) &
            (hourly["DATETIME"] < end)
        ]

        if hourly.empty:
            logger.warning("No data found for %s", self.target_date.date())
            return None

        # ------------------------------------------------
        # STEP 4: Material-wise aggregation (hour-specific)
        # ------------------------------------------------
        final_rows = []

        for _, row in hourly.iterrows():
            hr = row["DATETIME"]

            # Fetch Neon mapping valid for THIS hour
            mapping = self._fetch_material_mapping_for_hour(hr)

            material_row = {"DATETIME": hr}

            for mat, hopper_list in mapping.items():
                present_hoppers = [h for h in hopper_list if h in row.index]
                material_row[mat] = (
                    row[present_hoppers].sum() / 1000
                    if present_hoppers else 0
                )

            final_rows.append(material_row)

        agg = pd.DataFrame(final_rows)

        # ------------------------------------------------
        # STEP 5: Grouped material totals
        # ------------------------------------------------
        for group_name, child_materials in self.material_groups.items():
            cols_present = [c for c in child_materials if c in agg.columns]
            agg[group_name] = agg[cols_present].sum(axis=1) if cols_present else 0

        # ------------------------------------------------
        # STEP 6: PCI enrichment
        # ------------------------------------------------
        pci_df = self._get_pci_mt_from_influx()
        if not pci_df.empty:
            agg = agg.merge(pci_df, on="DATETIME", how="left")
            logger.info("pci_mt data merged successfully.")
        if "pci_mt" not in agg.columns:
            agg["pci_mt"] = 0

        # ------------------------------------------------
        # STEP 7: Column ordering & rename
        # ------------------------------------------------
        ordered = (
            ["DATETIME"] +
            [c for c in agg.columns if c not in ("DATETIME", "pci_mt")] +
            ["pci_mt"]
        )

        agg = agg[ordered]

        rename_map = config.get("charge_fields", {})
        agg = agg.rename(columns=rename_map)

        # ------------------------------------------------
        # STEP 8: Save output
        # ------------------------------------------------
        out = os.path.join(
            self.output_dir,
            f"hourly_materials_{self.target_date:%Y_%m_%d}.xlsx"
        )
        agg.to_excel(out, index=False)
        logger.info("Hourly material report saved: %s", out)

        return agg







def process_rm_hm_sheet(file_path: str, config: dict, start_date: list, output_dir: str = None) -> pd.DataFrame:
    """
    Reads the RM & HM combined Excel sheet (e.g., 'SP-02'),
    cleans column names, filters by date(s), and fills missing
    values (ai, ti, rdi, ri) with last available values from previous days.
    """
    try:
        sheet_name = config.get("rm_hm", {}).get("sheet_name", "SP-02")

        # --- Read sheet ---
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

        # --- Cleanup columns ---
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.replace(r"\s+", "_", regex=True)
            .str.replace(r"[^0-9a-zA-Z_]", "", regex=True)
            .str.lower()
        )

        df = df.dropna(how="all")

        # --- Standardize date column ---
        for col in df.columns:
            if col.lower() in ["date", "dates", "dt"]:
                df = df.rename(columns={col: "date"})
                break

        if "date" not in df.columns:
            logging.warning(f"'date' column not found. Available columns: {list(df.columns)}")
            return pd.DataFrame()

        # --- Convert date properly ---
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.sort_values("date").reset_index(drop=True)

        # --- Define the key columns to fill ---
        target_cols = ["ai", "ti", "rdi", "ri"]
        for col in target_cols:
            if col not in df.columns:
                logging.warning(f"Column '{col}' not found in sheet.")
                df[col] = None

        # --- Forward fill to propagate last known values ---
        df[target_cols] = df[target_cols].ffill()

        # --- Filter only the requested date(s) ---
        target_dates = [pd.to_datetime(d, format="%d-%b-%Y") for d in start_date]
        filtered_df = df[df["date"].isin(target_dates)].copy()

        # --- Format date back to string for consistency ---
        filtered_df["date"] = filtered_df["date"].dt.strftime("%d-%b-%Y")

        # --- Save optionally ---
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            out_path = os.path.join(output_dir, f"filtered_rm_hm_{sheet_name}.xlsx")
            filtered_df.to_excel(out_path, index=False)

        return filtered_df

    except Exception as e:
        logging.error(f"Error reading RM & HM sheet from '{file_path}': {e}")
        return pd.DataFrame()
